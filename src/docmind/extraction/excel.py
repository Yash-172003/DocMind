"""Excel (.xlsx) extraction via openpyxl.

A spreadsheet is structurally different from a PDF or a Word document:
there's no "prose" to extract, only a grid of cells per sheet. We treat
each sheet as both one ExtractedPage (its cells rendered as
pipe-delimited text, so it's still searchable/embeddable as text later)
and one ExtractedTable (the same cells, kept as a proper grid for callers
that need structure instead of flattened text).
"""

from io import BytesIO

from openpyxl import load_workbook

from docmind.extraction.exceptions import CorruptDocumentError
from docmind.extraction.models import ExtractedPage, ExtractedTable, ExtractionResult


def extract_xlsx(data: bytes) -> ExtractionResult:
    """Extract every sheet's cells from .xlsx bytes."""
    try:
        workbook = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise CorruptDocumentError(f"Could not open Excel workbook: {exc}") from exc

    pages: list[ExtractedPage] = []
    tables: list[ExtractedTable] = []
    warnings: list[str] = []

    for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
        sheet = workbook[sheet_name]
        rows: list[list[str | None]] = [
            [str(cell) if cell is not None else None for cell in row]
            for row in sheet.iter_rows(values_only=True)
        ]

        if not any(any(cell is not None for cell in row) for row in rows):
            warnings.append(f"sheet '{sheet_name}': empty, no cells found")

        text = "\n".join(
            " | ".join(cell or "" for cell in row) for row in rows if any(row)
        )
        pages.append(ExtractedPage(page_number=sheet_index, text=text))
        tables.append(ExtractedTable(page_number=sheet_index, rows=rows))

    workbook.close()

    full_text = "\n\n".join(p.text for p in pages if p.text)
    metadata: dict[str, str | int | float | bool | None] = {
        "sheet_count": len(pages),
        "sheet_names": ", ".join(workbook.sheetnames),
    }

    return ExtractionResult(
        text=full_text,
        pages=pages,
        tables=tables,
        metadata=metadata,
        warnings=warnings,
    )
